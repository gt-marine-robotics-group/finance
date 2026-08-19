"""
order_excel_builder.py - Side-by-side Budget vs Quoted Full Detail report generator.

Generates beautifully formatted Excel (.xlsx) and CSV (.csv) comparison reports matching the
PR26968_Budget_vs_Quoted_Full_Detail standard format for attachment to Engage purchase requests.
"""

from __future__ import annotations

import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_order_budget_vs_quoted_excel(
    order_id: str,
    requests_to_submit: list[dict],
    bill_line_cache: dict = None,
    scraped_results: dict = None,
    output_dir: str = None
) -> tuple[str, str]:
    """
    Generate side-by-side Budget vs Quoted Excel (.xlsx) and CSV (.csv) comparison reports.
    Returns (xlsx_path, csv_path).
    """
    if not output_dir:
        output_dir = os.path.join("screenshots", order_id)
    os.makedirs(output_dir, exist_ok=True)

    # Group items by bill_no
    bills_grouped = {}
    for r in requests_to_submit:
        b_no = str(r.get("bill_no") or "376851").strip()
        bills_grouped.setdefault(b_no, []).append(r)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget vs Quoted Detail"

    # Styling definitions
    font_title = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="595959")
    font_section_hdr = Font(name="Calibri", size=11, bold=True, color="1F4E78")
    font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)

    fill_hdr = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    total_border = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000")
    )

    # Title Block
    all_bill_nos = " & ".join(f"Bill {b}" for b in bills_grouped.keys())
    ws.cell(row=1, column=1, value="Budget Request & Quoted Line Items Comparison").font = font_title
    ws.cell(row=2, column=1, value=f"Complete side-by-side mapping of Budget Request items with all Quoted Bill Line Items ({all_bill_nos})").font = font_subtitle

    # Table Headers (Row 4)
    headers = [
        "Budget Line #", "Quoted Bill #", "Budget Item Description", "Category",
        "Budget Qty", "Budget Unit Cost", "Budget Total",
        "Quoted Quantity", "Quoted Unit Cost", "Quoted Total", "Variance (Quoted - Budget)"
    ]

    header_row_num = 4
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_num, column=col_idx, value=h)
        cell.font = font_hdr
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    curr_row = 5
    global_line_counter = 1
    total_budget_grand = 0.0
    total_quoted_grand = 0.0

    sub_row_indices = []
    bill_keys = list(bills_grouped.keys())

    for b_idx, sec_bill in enumerate(bill_keys):
        sec_items = bills_grouped.get(sec_bill, [])
        if not sec_items:
            continue

        if b_idx > 0:
            curr_row += 1
            sec_title = f"Additional Quoted Line Items (Bill {sec_bill})"
            cell_sec = ws.cell(row=curr_row, column=1, value=sec_title)
            cell_sec.font = font_section_hdr
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=11)
            curr_row += 1

        sec_start_row = curr_row

        for r in sec_items:
            item_name = r.get("item_name", "")
            sec_cache = bill_line_cache.get(sec_bill) or {} if bill_line_cache else {}
            loc = sec_cache.get(item_name)
            if not loc and sec_cache:
                import engage_bill_lookup
                loc = engage_bill_lookup.find_best_item_match(item_name, sec_cache)
            sec_line = loc.get("section_line_number") if loc else None
            line_str = f"Line {sec_line or global_line_counter}"
            global_line_counter += 1

            sec_name = str(loc.get("section") or r.get("source_bill_title") or "B03 - General Inventoried Goods").strip() if loc else "B03 - General Inventoried Goods"
            qty = int(r.get("quantity", 1))
            alloc_cost = float(r.get("cost", 0.0))

            live_val = scraped_results.get(item_name) if scraped_results else None
            quoted_cost = float(live_val) if live_val is not None else alloc_cost

            r_idx = curr_row

            # Use live Excel formulas for totals and variance
            formula_alloc_total = f"=E{r_idx}*F{r_idx}"
            formula_quoted_total = f"=H{r_idx}*I{r_idx}"
            formula_variance = f"=J{r_idx}-G{r_idx}"

            row_vals = [
                line_str, f"Bill {sec_bill}", item_name, sec_name,
                qty, alloc_cost, formula_alloc_total,
                qty, quoted_cost, formula_quoted_total,
                formula_variance
            ]

            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=curr_row, column=col_idx, value=val)
                cell.font = font_regular
                cell.border = thin_border
                if curr_row % 2 == 1:
                    cell.fill = fill_zebra
                if col_idx in (6, 7, 9, 10, 11):
                    cell.number_format = "$#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in (1, 2, 5, 8):
                    cell.alignment = Alignment(horizontal="center")

            curr_row += 1

        sec_end_row = curr_row - 1

        # Subtotal Row per Bill with live Excel SUM formulas
        sub_row_idx = curr_row
        sub_row_indices.append(sub_row_idx)

        formula_sub_budget = f"=SUM(G{sec_start_row}:G{sec_end_row})"
        formula_sub_quoted = f"=SUM(J{sec_start_row}:J{sec_end_row})"
        formula_sub_variance = f"=J{sub_row_idx}-G{sub_row_idx}"

        sub_row_vals = ["", "", "", "", "", f"Subtotal (Bill {sec_bill}):", formula_sub_budget, "", "", formula_sub_quoted, formula_sub_variance]
        for col_idx, val in enumerate(sub_row_vals, start=1):
            cell = ws.cell(row=curr_row, column=col_idx, value=val)
            cell.font = font_bold
            cell.border = total_border
            if col_idx in (6, 7, 9, 10, 11):
                if isinstance(val, str) and val.startswith("="):
                    cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        curr_row += 1

    # Grand Total Row: Sum subtotal rows directly to avoid double-counting subtotal rows
    curr_row += 1
    grand_row_idx = curr_row

    if len(sub_row_indices) == 1:
        formula_grand_budget = f"=G{sub_row_indices[0]}"
        formula_grand_quoted = f"=J{sub_row_indices[0]}"
    elif len(sub_row_indices) > 1:
        b_refs = ", ".join(f"G{r}" for r in sub_row_indices)
        q_refs = ", ".join(f"J{r}" for r in sub_row_indices)
        formula_grand_budget = f"=SUM({b_refs})"
        formula_grand_quoted = f"=SUM({q_refs})"
    else:
        formula_grand_budget = "=0"
        formula_grand_quoted = "=0"

    formula_grand_variance = f"=J{grand_row_idx}-G{grand_row_idx}"

    grand_row_vals = ["", "", "", "", "", "Grand Total:", formula_grand_budget, "", "", formula_grand_quoted, formula_grand_variance]
    for col_idx, val in enumerate(grand_row_vals, start=1):
        cell = ws.cell(row=curr_row, column=col_idx, value=val)
        cell.font = font_bold
        cell.border = total_border
        if col_idx in (6, 7, 9, 10, 11):
            if isinstance(val, str) and val.startswith("="):
                cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    # Auto-adjust column widths (skip title rows 1-3 when measuring text lengths)
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter == "A":
            ws.column_dimensions["A"].width = 16
        else:
            # Measure text length starting from row 4 (table header), ignoring long merged title lines
            lengths = [len(str(cell.value or "")) for cell in col[3:] if "Additional Quoted" not in str(cell.value or "")]
            max_len = max(lengths) if lengths else 12
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    xlsx_path = os.path.join(output_dir, f"Budget_vs_Quoted_Detail_{order_id}.xlsx")
    csv_path = os.path.join(output_dir, f"Budget_vs_Quoted_Detail_{order_id}.csv")

    wb.save(xlsx_path)

    # Save CSV version
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([v if v is not None else "" for v in row])

    return xlsx_path, csv_path


def main():
    import argparse
    import pandas as pd
    import spreadsheet_utils
    import price_scraper

    parser = argparse.ArgumentParser(description="Generate Budget vs Quoted Full Detail Excel and CSV comparison reports.")
    parser.add_argument("--order", required=True, help="Order ID (e.g. 260811_amazon_awu335)")
    parser.add_argument("--excel-path", default="FY27_Bills_Budget.xlsx", help="Path to master Excel file")
    parser.add_argument("--skip-scrape", action="store_true", help="Skip live web price scraping and use spreadsheet allocations")
    args = parser.parse_args()

    order_id = args.order
    excel_path = args.excel_path
    cwd_path = os.path.join(os.getcwd(), "FY27_Bills_Budget.xlsx")
    repo_path = os.path.expanduser("~/mrg/finance/FY27_Bills_Budget.xlsx")

    if os.path.exists(cwd_path):
        excel_path = cwd_path
    elif os.path.exists(repo_path):
        excel_path = repo_path
    elif not os.path.exists(excel_path):
        print(f"❌ Master spreadsheet not found at {excel_path}")
        return

    wb_in = openpyxl.load_workbook(excel_path, data_only=True)
    df_bills = spreadsheet_utils.read_sheet_robust(wb_in, ["Bills", "Bill", "Budget"])
    bill_item_map = {spreadsheet_utils.get_col_val(r.to_dict(), "bill_item_id"): r.to_dict() for _, r in df_bills.iterrows() if spreadsheet_utils.get_col_val(r.to_dict(), "bill_item_id")}

    df_orders = spreadsheet_utils.read_sheet_robust(wb_in, ["Ordering", "Orders", "OrderT"])
    oid_col = next((c for c in df_orders.columns if "order" in str(c).lower()), "Order ID")

    order_rows = [r for _, r in df_orders.iterrows() if str(r.get(oid_col, "")).strip() == order_id]
    if not order_rows:
        print(f"❌ Order '{order_id}' not found in {excel_path}")
        return

    requests_to_submit = []
    scraped_results = {}
    print(f"🔍 Loading {len(order_rows)} item(s) for order {order_id}...")
    for i, row in enumerate(order_rows, 1):
        b_id = spreadsheet_utils.get_col_val(row.to_dict(), "bill_item_id")
        b_row = bill_item_map.get(b_id, {})
        b_no = spreadsheet_utils.get_col_val(b_row, "bill_no") or spreadsheet_utils.get_col_val(row.to_dict(), "bill_no") or "376851"
        item_name = spreadsheet_utils.get_col_val(row.to_dict(), "item_name") or spreadsheet_utils.get_col_val(b_row, "item_name")
        sec = spreadsheet_utils.get_col_val(b_row, "budget_section") or "B03 - General Inventoried Goods"
        link = spreadsheet_utils.get_col_val(row.to_dict(), "link") or spreadsheet_utils.get_col_val(b_row, "link")
        cost = spreadsheet_utils.safe_float(b_row.get("Cost", row.to_dict().get("Allocation", 0)))
        qty = spreadsheet_utils.safe_int(row.to_dict().get("Quantity", 1))

        if not args.skip_scrape and link and link.startswith("http"):
            res = price_scraper.scrape_item_price(link)
            if res and res.get("current_price") is not None:
                scraped_results[item_name] = float(res["current_price"])
                if float(res["current_price"]) != cost:
                    print(f"  💰 Live price scraped for '{item_name}': ${float(res['current_price']):.2f} (Allocated: ${cost:.2f})")

        requests_to_submit.append({
            "item_name": item_name,
            "quantity": qty,
            "cost": cost,
            "total": cost * qty,
            "bill_no": b_no,
            "bill_item_id": b_id,
            "link": link
        })

    xlsx_path, csv_path = generate_order_budget_vs_quoted_excel(
        order_id=order_id,
        requests_to_submit=requests_to_submit,
        scraped_results=scraped_results
    )

    print(f"\n✅ Report Generation Complete!")
    print(f"  📗 Excel Spreadsheet: {xlsx_path}")
    print(f"  📄 CSV Spreadsheet:   {csv_path}")


if __name__ == "__main__":
    main()
