"""
test_order_excel_builder.py - Unit tests for dynamic Budget vs Quoted Excel report generation.
"""

import os
import openpyxl
import pytest
import order_excel_builder


def test_single_bill_various_lengths(tmp_path):
    """Test generating report for a single bill with 3 items."""
    items = [
        {"item_name": "Item A", "quantity": 2, "cost": 10.0, "bill_no": "1001"},
        {"item_name": "Item B", "quantity": 1, "cost": 25.0, "bill_no": "1001"},
        {"item_name": "Item C", "quantity": 5, "cost": 4.0,  "bill_no": "1001"},
    ]
    out_dir = str(tmp_path)
    xlsx_path, csv_path = order_excel_builder.generate_order_budget_vs_quoted_excel(
        order_id="test_single_bill",
        requests_to_submit=items,
        output_dir=out_dir
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Item rows 5, 6, 7. Subtotal row 8. Grand total row 10.
    assert ws.cell(row=5, column=7).value == "=E5*F5"
    assert ws.cell(row=7, column=7).value == "=E7*F7"

    # Subtotal formula should be =SUM(G5:G7)
    assert ws.cell(row=8, column=7).value == "=SUM(G5:G7)"
    assert ws.cell(row=8, column=10).value == "=SUM(J5:J7)"

    # Grand total formula should reference subtotal row G8
    assert ws.cell(row=10, column=7).value == "=G8"
    assert ws.cell(row=10, column=10).value == "=J8"


def test_multi_bill_different_lengths(tmp_path):
    """Test generating report for multi-bill order with different item counts per bill."""
    items = [
        # Bill 1001 (2 items)
        {"item_name": "Resistor Pack", "quantity": 10, "cost": 1.5, "bill_no": "1001"},
        {"item_name": "Capacitor Pack", "quantity": 5,  "cost": 3.0, "bill_no": "1001"},
        # Bill 2002 (4 items)
        {"item_name": "Sensor 1", "quantity": 1, "cost": 50.0, "bill_no": "2002"},
        {"item_name": "Sensor 2", "quantity": 1, "cost": 45.0, "bill_no": "2002"},
        {"item_name": "Sensor 3", "quantity": 2, "cost": 20.0, "bill_no": "2002"},
        {"item_name": "Sensor 4", "quantity": 1, "cost": 30.0, "bill_no": "2002"},
    ]
    out_dir = str(tmp_path)
    xlsx_path, csv_path = order_excel_builder.generate_order_budget_vs_quoted_excel(
        order_id="test_multi_bill",
        requests_to_submit=items,
        output_dir=out_dir
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Bill 1001 items: rows 5, 6 -> Subtotal row 7
    assert ws.cell(row=7, column=7).value == "=SUM(G5:G6)"

    # Bill 2002 header row 9, items: rows 10, 11, 12, 13 -> Subtotal row 14
    assert ws.cell(row=14, column=7).value == "=SUM(G10:G13)"

    # Grand Total row 16 should reference subtotal rows G7 and G14
    assert ws.cell(row=16, column=7).value == "=SUM(G7, G14)"
    assert ws.cell(row=16, column=10).value == "=SUM(J7, J14)"
    assert ws.cell(row=16, column=11).value == "=J16-G16"
