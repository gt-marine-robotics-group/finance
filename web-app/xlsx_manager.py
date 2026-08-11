"""
xlsx_manager.py - Read/write FY27_Bills_Budget.xlsx with rclone sync.

Pull before read, push after write. The xlsx on SharePoint is the source of truth.
"""

import os
import sys
import subprocess
import threading
from pathlib import Path
from openpyxl import load_workbook

# Add parent directory for price_scraper import
parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

import price_scraper

# Config from environment
RCLONE_REMOTE = os.environ.get(
    "RCLONE_REMOTE",
    "onedrive:OPS-1 Operations/FY27 Finances/FY27_Bills_Budget.xlsx",
)
LOCAL_XLSX = os.environ.get("LOCAL_XLSX_PATH", os.path.expanduser("~/mrg/finance/FY27_Bills_Budget.xlsx"))
SHEET_NAME = os.environ.get("XLSX_SHEET_NAME", "Bills")
QUEUE_SHEET_NAME = os.environ.get("XLSX_QUEUE_SHEET_NAME", "Test")

# File lock to prevent concurrent reads/writes
_lock = threading.Lock()

# Cache: only pull from SharePoint every N seconds
_last_pull_time = 0
PULL_INTERVAL = int(os.environ.get("PULL_INTERVAL_SECONDS", "300"))  # 5 minutes

# Cached data
_cached_items: list[dict] = []
_cached_items_time = 0
_cached_queue: list[dict] = []
_cached_queue_time = 0
ITEMS_CACHE_TTL = 300  # 5 minutes — hit Sync to force refresh


def invalidate_items_cache():
    """Invalidate items cache and reset pull timer."""
    global _cached_items, _cached_items_time, _last_pull_time
    _cached_items = []
    _cached_items_time = 0
    _last_pull_time = 0


def invalidate_queue_cache():
    """Invalidate test queue cache."""
    global _cached_queue, _cached_queue_time
    _cached_queue = []
    _cached_queue_time = 0


def invalidate_all_caches():
    """Invalidate all cached items and queue data."""
    invalidate_items_cache()
    invalidate_queue_cache()


# Column mapping (xlsx columns in the Bills sheet)
COLUMNS = [
    "Bill Item ID",
    "Bill No.",
    "Bill Title",
    "Item Name",
    "Status",
    "Budget Section",
    "Vendor",
    "Description",
    "Quantity",
    "Cost",
    "Total Cost",
    "Link",
    "File URL",
    "Person Requesting",
]


def _run_rclone(args: list[str]) -> bool:
    """Run an rclone command. Returns True on success."""
    try:
        result = subprocess.run(
            ["rclone"] + args,
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            print(f"[rclone] Error: {result.stderr.strip()}")
            return False
        return True
    except FileNotFoundError:
        print("[rclone] rclone not found - skipping sync")
        return False
    except subprocess.TimeoutExpired:
        print("[rclone] Timeout during sync")
        return False


def sync_pull() -> bool:
    """Pull latest xlsx from SharePoint. Skips if pulled recently (within PULL_INTERVAL)."""
    global _last_pull_time
    import time as _time

    now = _time.time()
    if now - _last_pull_time < PULL_INTERVAL:
        return True  # Use cached local copy

    local_dir = str(Path(LOCAL_XLSX).parent)
    os.makedirs(local_dir, exist_ok=True)
    result = _run_rclone(["copy", "--checksum", RCLONE_REMOTE, local_dir])
    if result:
        _last_pull_time = now
    return result


def _get_graph_token() -> tuple[str, str, str] | None:
    """Read access token, drive_id, and file_id from rclone config / cache."""
    import configparser
    import json as _json

    rclone_conf = os.path.expanduser("~/.config/rclone/rclone.conf")
    if not os.path.exists(rclone_conf):
        return None

    config = configparser.ConfigParser()
    config.read(rclone_conf)

    if "onedrive" not in config:
        return None

    try:
        token_str = config["onedrive"]["token"]
        token = _json.loads(token_str)
        drive_id = config["onedrive"]["drive_id"]
        access_token = token["access_token"]
    except (KeyError, _json.JSONDecodeError):
        return None

    # Get file ID (cache it after first lookup)
    file_id = os.environ.get("_GRAPH_FILE_ID", "")
    if not file_id:
        import requests as _requests
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/OPS-1 Operations/FY27 Finances/FY27_Bills_Budget.xlsx"
        resp = _requests.get(url, headers={"Authorization": f"Bearer {access_token}"}, timeout=10)
        if resp.status_code == 200:
            file_id = resp.json()["id"]
            os.environ["_GRAPH_FILE_ID"] = file_id
        else:
            return None

    return access_token, drive_id, file_id


def graph_add_row(sheet_table: str, row_values: list, index: int | None = None) -> bool:
    """Add a row to a table via Graph API Excel workbook endpoint.
    If index is provided, inserts at that position. Otherwise appends to end.
    """
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        print("[graph] No credentials - skipping")
        return False

    access_token, drive_id, file_id = creds

    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/{sheet_table}/rows/add"
    payload = {"values": [row_values]}
    if index is not None:
        payload["index"] = index

    resp = _requests.post(
        url,
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=15,
    )

    if resp.status_code in (200, 201):
        print(f"[graph] ✅ Row added to {sheet_table}" + (f" at index {index}" if index is not None else ""))
        return True
    else:
        print(f"[graph] ❌ Failed to add row: {resp.status_code} {resp.text[:150]}")
        return False


def _get_last_data_index(sheet_table: str) -> int:
    """Find the index of the last non-empty row in a table."""
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        return -1

    access_token, drive_id, file_id = creds
    headers = {"Authorization": f"Bearer {access_token}"}

    rows_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/{sheet_table}/rows"
    resp = _requests.get(rows_url, headers=headers, timeout=15)

    if resp.status_code != 200:
        return -1

    last_data_index = -1
    for row in resp.json().get("value", []):
        vals = row["values"][0] if row.get("values") else []
        if any(str(v).strip() for v in vals if v):
            last_data_index = row["index"]

    return last_data_index


def _patch_row_values(sheet_table: str, row_index: int, values: list, columns: list, skip_columns: set, headers: dict, drive_id: str, file_id: str) -> bool:
    """
    Write values to specific cells in an existing row, SKIPPING formula columns entirely.
    Uses cell-level updates so formulas in other columns are never touched.
    """
    import requests as _requests

    # Table data starts at sheet row 2 (row 1 = header)
    sheet_row = row_index + 2

    # Map column names to Excel column letters
    col_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T']

    # Build updates for non-formula columns only
    cells_to_update = {}
    for i, col in enumerate(columns):
        if col in skip_columns:
            continue  # Don't touch formula columns at all
        if i < len(values) and i < len(col_letters):
            val = values[i]
            if val or val == 0:  # Write value (including 0)
                cells_to_update[f"{col_letters[i]}{sheet_row}"] = val

    # Batch update using range PATCH for each cell
    for cell_addr, val in cells_to_update.items():
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/worksheets('Bills')/range(address='{cell_addr}')"
        resp = _requests.patch(url, headers=headers, json={"values": [[val]]}, timeout=10)
        if resp.status_code != 200:
            print(f"[graph] ⚠️ Failed to write {cell_addr}: {resp.status_code}")
            return False

    print(f"[graph] ✅ Wrote {len(cells_to_update)} cells to row {sheet_row}")
    return True


def graph_get_table_columns(sheet_table: str) -> list[str]:
    """Get column names for a table via Graph API."""
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        return []

    access_token, drive_id, file_id = creds

    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/{sheet_table}/columns"
    resp = _requests.get(
        url,
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=10,
    )

    if resp.status_code == 200:
        return [c["name"] for c in resp.json()["value"]]
    return []


def sync_push() -> bool:
    """No-op — writes now go directly via Graph API."""
    return True


def _push_async():
    """No-op — writes are immediate via Graph API."""
    pass


# Titles to skip (same as automation.py)
SKIP_TITLE_PREFIXES = ("nan", "request", "liquid", "misc")


def read_items() -> list[dict]:
    """
    Read all items from the Bills sheet. Cached for 60s.
    """
    global _cached_items, _cached_items_time
    import time as _time

    now = _time.time()
    if _cached_items and (now - _cached_items_time < ITEMS_CACHE_TTL):
        return _cached_items

    items = _read_items_from_xlsx()
    _cached_items = items
    _cached_items_time = now
    return items


def _read_items_from_xlsx() -> list[dict]:
    """
    Read all items from the Bills sheet (actual read).
    Pulls latest from SharePoint first.
    Auto-detects header row by scanning for 'Item Name' column.
    Skips non-bill items (Liquid, Misc, etc.) same as automation.py.
    Returns list of dicts with normalized keys.
    """
    try:
        with _lock:
            sync_pull()

        if not os.path.exists(LOCAL_XLSX):
            return []

        wb = load_workbook(LOCAL_XLSX, read_only=True, data_only=True)
        ws = wb[SHEET_NAME]

        # Auto-detect header row by finding the row containing "Item Name"
        header_row = None
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_values = [str(cell).strip() if cell else "" for cell in row]
            if "Item Name" in row_values:
                header_row = row_idx
                headers = row_values
                break

        if header_row is None:
            wb.close()
            return []

        items = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(row):
                continue
            item = {}
            for i, col in enumerate(headers):
                if not col:
                    continue
                if i < len(row):
                    val = row[i]
                    # Normalize whitespace in strings
                    if isinstance(val, str):
                        val = " ".join(val.split())
                    item[col] = val if val is not None else ""
                else:
                    item[col] = ""
            # Only include rows that have an Item Name
            if not item.get("Item Name"):
                continue
            # Skip non-bill items (same filter as automation.py)
            bill_title = str(item.get("Bill Title", "")).strip().lower()
            if any(bill_title.startswith(prefix) for prefix in SKIP_TITLE_PREFIXES):
                continue
            # Skip items with negative Bill Item IDs (metadata rows)
            try:
                item_id = float(str(item.get("Bill Item ID", 0)))
                if item_id < 0:
                    continue
            except (ValueError, TypeError):
                pass
            items.append(item)

        wb.close()
        return items
    except Exception as e:
        print(f"[read_items] Error: {e}")
        return []


def get_bills() -> list[str]:
    """Get list of unique bill titles (non-empty)."""
    items = read_items()
    titles = set()
    for item in items:
        title = str(item.get("Bill Title", "")).strip()
        if title:
            titles.add(title)
    return sorted(titles)


def get_items_by_bill(bill_title: str) -> list[dict]:
    """Get items for a specific bill."""
    items = read_items()
    return [i for i in items if str(i.get("Bill Title", "")).strip() == bill_title]


def get_backlog_items() -> list[dict]:
    """Get items from the queue (Test sheet)."""
    return read_queue_items()


def _find_row_by_item_id(ws, item_id: str, headers: list[str], header_row: int) -> int | None:
    """Find the row number for a given Bill Item ID."""
    id_col = None
    for i, h in enumerate(headers):
        if h == "Bill Item ID":
            id_col = i
            break
    if id_col is None:
        return None

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if id_col < len(row) and str(row[id_col]).strip() == str(item_id).strip():
            return row_idx
    return None


def _get_headers(ws) -> tuple[list[str], int]:
    """Get header row as list of strings and the 1-indexed row number."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_values = [str(cell).strip() if cell else "" for cell in row]
        if "Item Name" in row_values:
            return row_values, row_idx
    # Fallback to row 1
    return [str(cell.value).strip() if cell.value else "" for cell in ws[1]], 1


def _get_next_item_id(ws, headers: list[str], header_row: int) -> int:
    """Get the next available Bill Item ID."""
    id_col = None
    for i, h in enumerate(headers):
        if h == "Bill Item ID":
            id_col = i
            break
    if id_col is None:
        return 1

    max_id = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if id_col < len(row) and row[id_col]:
            try:
                val = int(float(str(row[id_col])))
                max_id = max(max_id, val)
            except (ValueError, TypeError):
                pass
    return max_id + 1


def add_item(item_data: dict) -> bool:
    """
    Add a new item to the QUEUE (TestTable on Test sheet) via Graph API.
    Also writes locally for immediate display.
    """
    if "Vendor" in item_data and item_data["Vendor"]:
        item_data["Vendor"] = price_scraper.normalize_vendor(item_data["Vendor"])

    # Get TestTable columns
    columns = graph_get_table_columns("TestTable")
    if not columns:
        # Fallback: use known columns
        columns = ["Bill Item ID", "Bill No.", "Bill Title", "Item Name",
                   "Budget Section", "Quantity", "Cost", "Vendor", "Description", "Link", "Column1"]

    # Build row values in column order
    row_values = []
    for col in columns:
        val = item_data.get(col, "")
        row_values.append(val if val else "")

    # Push to SharePoint via Graph API
    success = graph_add_row("TestTable", row_values)

    # Invalidate queue cache so next read gets fresh data with correct _table_index
    if success:
        invalidate_queue_cache()

    # Also write locally for immediate read-back
    with _lock:
        if os.path.exists(LOCAL_XLSX):
            try:
                wb = load_workbook(LOCAL_XLSX)
                ws = wb[QUEUE_SHEET_NAME]
                headers, header_row = _get_headers(ws)
                local_row = []
                for h in headers:
                    local_row.append(item_data.get(h, ""))
                ws.append(local_row)
                wb.save(LOCAL_XLSX)
                wb.close()
            except Exception as e:
                print(f"[local] Write failed: {e}")

    return success



def read_queue_items() -> list[dict]:
    """
    Read all items from the Queue (TestTable) via Graph API.
    Cached for ITEMS_CACHE_TTL seconds.
    Falls back to local xlsx if Graph API unavailable.
    """
    global _cached_queue, _cached_queue_time
    import time as _time
    import requests as _requests

    now = _time.time()
    if _cached_queue and (now - _cached_queue_time < ITEMS_CACHE_TTL):
        return _cached_queue

    result = _fetch_queue_items()
    _cached_queue = result
    _cached_queue_time = now
    return result


def _fetch_queue_items() -> list[dict]:
    """Fetch queue items from Graph API or local xlsx."""
    import requests as _requests

    creds = _get_graph_token()
    if creds:
        access_token, drive_id, file_id = creds
        headers = {"Authorization": f"Bearer {access_token}"}

        # Get columns
        cols_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/TestTable/columns"
        cols_resp = _requests.get(cols_url, headers=headers, timeout=10)

        # Get rows
        rows_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/TestTable/rows"
        rows_resp = _requests.get(rows_url, headers=headers, timeout=10)

        if cols_resp.status_code == 200 and rows_resp.status_code == 200:
            columns = [c["name"] for c in cols_resp.json()["value"]]
            items = []
            for row in rows_resp.json().get("value", []):
                values = row["values"][0] if row.get("values") else []
                item = {}
                for i, col in enumerate(columns):
                    if i < len(values):
                        val = values[i]
                        if isinstance(val, str):
                            val = " ".join(val.split())
                        item[col] = val if val else ""
                    else:
                        item[col] = ""
                if item.get("Item Name"):
                    item["_table_index"] = row["index"]
                    items.append(item)
            return items

    # Fallback to local xlsx
    with _lock:
        sync_pull()
        if not os.path.exists(LOCAL_XLSX):
            return []

        wb = load_workbook(LOCAL_XLSX, read_only=True, data_only=True)
        ws = wb[QUEUE_SHEET_NAME]
        headers_list, header_row = _get_headers(ws)

        items = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(row):
                continue
            item = {}
            for i, col in enumerate(headers_list):
                if not col:
                    continue
                if i < len(row):
                    val = row[i]
                    if isinstance(val, str):
                        val = " ".join(val.split())
                    item[col] = val if val is not None else ""
                else:
                    item[col] = ""
            if item.get("Item Name"):
                item["_table_index"] = row_idx - header_row - 1  # 0-indexed for Graph API
                items.append(item)

        wb.close()
        return items


def delete_queue_item(row_idx: int) -> bool:
    """Delete an item from the queue (Test sheet) by row index."""
    with _lock:
        if not os.path.exists(LOCAL_XLSX):
            return False

        wb = load_workbook(LOCAL_XLSX)
        ws = wb[QUEUE_SHEET_NAME]
        ws.delete_rows(row_idx)
        wb.save(LOCAL_XLSX)
        wb.close()
        return True


def move_to_bill(queue_items: list[dict], bill_title: str, add_separator: bool = True, person: str = "") -> int:
    """
    Move items from the Queue (TestTable) to the Bills table (BillsT) via Graph API.
    - Inserts a separator row (e.g. "Request 4") before the items
    - Adds each item as a new row on BillsT with the given Bill Title
    - Deletes them from TestTable
    - Returns number of items successfully moved.
    """
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        print("[graph] No credentials")
        return 0

    access_token, drive_id, file_id = creds
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    # Get BillsT columns
    bills_columns = graph_get_table_columns("BillsT")
    if not bills_columns:
        bills_columns = ["Bill Item ID", "Bill No.", "Bill Title", "Item Name", "Status",
                         "Budget Section", "Vendor", "Description", "Quantity", "Cost",
                         "Total Cost", "Link", "File URL", "Person Requesting", "Remaining Allocation", "Column1"]

    # Get existing rows to find next Request number
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/BillsT/rows"
    resp = _requests.get(url, headers=headers, timeout=15)
    max_request_num = 0
    if resp.status_code == 200:
        for row in resp.json().get("value", []):
            vals = row["values"][0]
            bill_col = str(vals[2]) if vals[2] else ""
            if bill_col.startswith("Request"):
                try:
                    num = int(bill_col.replace("Request", "").strip())
                    if num > max_request_num:
                        max_request_num = num
                except ValueError:
                    pass

    next_request = max_request_num + 1

    # Find the first empty row to write to (formulas already exist there)
    insert_at = _get_last_data_index("BillsT") + 1

    # Columns with formulas — don't overwrite these
    FORMULA_COLUMNS = {"Bill Item ID", "Total Cost"}

    # Reserve the separator row slot if needed
    separator_row_idx = None
    if add_separator:
        separator_row_idx = insert_at
        insert_at += 1  # Items start after the separator slot

    moved = 0
    rows_to_delete = []

    for item in queue_items:
        item_data = dict(item)
        item_data["Bill Title"] = bill_title
        item_data["Status"] = "bill requested"
        if person:
            item_data["Person Requesting"] = person

        row_values = []
        for col in bills_columns:
            val = item_data.get(col, "")
            row_values.append(val if val else "")

        # PATCH existing empty row — preserves formulas in Bill Item ID and Total Cost
        success = _patch_row_values("BillsT", insert_at, row_values, bills_columns, FORMULA_COLUMNS, headers, drive_id, file_id)
        if success:
            moved += 1
            insert_at += 1
            if "_table_index" in item:
                rows_to_delete.append(item["_table_index"])

    # Write separator row ONLY if items were successfully added
    if moved > 0 and separator_row_idx is not None:
        sep_values = [""] * len(bills_columns)
        bill_title_idx = bills_columns.index("Bill Title") if "Bill Title" in bills_columns else 2
        sep_values[bill_title_idx] = f"Request {next_request}"
        _patch_row_values("BillsT", separator_row_idx, sep_values, bills_columns, FORMULA_COLUMNS, headers, drive_id, file_id)

    # Delete from TestTable (in reverse order so indices don't shift)
    rows_to_delete.sort(reverse=True)
    for idx in rows_to_delete:
        del_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/TestTable/rows/itemAt(index={idx})"
        resp = _requests.delete(del_url, headers=headers, timeout=10)
        if resp.status_code == 204:
            print(f"[graph] Deleted queue row index {idx}")
        else:
            print(f"[graph] Failed to delete row {idx}: {resp.status_code}")

    # Also sync local file
    sync_pull()

    return moved


def update_item(item_id: str, updates: dict) -> bool:
    """
    Update an existing item by Bill Item ID via Graph API or local fallback.
    Only modifies specified fields, preserving formula columns (Bill Item ID, Total Cost).
    """
    if not item_id:
        return False

    if "Vendor" in updates and updates["Vendor"]:
        updates["Vendor"] = price_scraper.normalize_vendor(updates["Vendor"])

    creds = _get_graph_token()
    if creds:
        access_token, drive_id, file_id = creds
        headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

        # Read rows to find row matching Bill Item ID
        rows_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/BillsT/rows"
        import requests as _requests
        resp = _requests.get(rows_url, headers=headers, timeout=15)

        if resp.status_code == 200:
            target_idx = None
            for row in resp.json().get("value", []):
                vals = row["values"][0] if row.get("values") else []
                if vals and str(vals[0]).strip() == str(item_id).strip():
                    target_idx = row["index"]
                    break

            if target_idx is not None:
                bills_columns = graph_get_table_columns("BillsT")
                sheet_row = target_idx + 2
                col_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T']

                FORMULA_COLUMNS = {"Bill Item ID", "Total Cost"}
                updated_count = 0
                for field_name, value in updates.items():
                    if field_name in FORMULA_COLUMNS:
                        continue
                    if field_name in bills_columns:
                        c_idx = bills_columns.index(field_name)
                        if c_idx < len(col_letters):
                            cell_addr = f"{col_letters[c_idx]}{sheet_row}"
                            url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/worksheets('Bills')/range(address='{cell_addr}')"
                            patch_resp = _requests.patch(url, headers=headers, json={"values": [[value]]}, timeout=10)
                            if patch_resp.status_code == 200:
                                updated_count += 1

                if updated_count > 0:
                    invalidate_items_cache()
                    return True

    # Fallback: update local xlsx
    with _lock:
        sync_pull()

        if not os.path.exists(LOCAL_XLSX):
            return False

        wb = load_workbook(LOCAL_XLSX)
        ws = wb[SHEET_NAME]
        headers, header_row = _get_headers(ws)

        row_idx = _find_row_by_item_id(ws, item_id, headers, header_row)
        if row_idx is None:
            wb.close()
            return False

        for key, value in updates.items():
            if key in headers and key not in ("Bill Item ID", "Total Cost"):
                col_idx = headers.index(key) + 1
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Recalculate Total Cost if qty or cost changed
        if "Quantity" in updates or "Cost" in updates:
            qty_col = headers.index("Quantity") + 1 if "Quantity" in headers else None
            cost_col = headers.index("Cost") + 1 if "Cost" in headers else None
            total_col = headers.index("Total Cost") + 1 if "Total Cost" in headers else None

            if qty_col and cost_col and total_col:
                qty = ws.cell(row=row_idx, column=qty_col).value or 1
                cost = ws.cell(row=row_idx, column=cost_col).value or 0
                try:
                    cost_val = float(str(cost).replace("$", "").replace(",", ""))
                    qty_val = float(qty)
                    ws.cell(row=row_idx, column=total_col, value=qty_val * cost_val)
                except (ValueError, TypeError):
                    pass

        wb.save(LOCAL_XLSX)
        wb.close()
        invalidate_items_cache()
        return True


def graph_get_order_rows() -> list[dict]:
    """Read all rows from OrderT via Graph API. Returns list of dicts keyed by column name."""
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        return []

    access_token, drive_id, file_id = creds
    headers = {"Authorization": f"Bearer {access_token}"}

    # Get columns
    cols_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/OrderT/columns"
    cols_resp = _requests.get(cols_url, headers=headers, timeout=10)

    # Get rows
    rows_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/OrderT/rows"
    rows_resp = _requests.get(rows_url, headers=headers, timeout=15)

    if cols_resp.status_code != 200 or rows_resp.status_code != 200:
        return []

    columns = [c["name"] for c in cols_resp.json()["value"]]
    items = []
    for row in rows_resp.json().get("value", []):
        values = row["values"][0] if row.get("values") else []
        item = {}
        for i, col in enumerate(columns):
            if i < len(values):
                val = values[i]
                if isinstance(val, str):
                    val = " ".join(val.split())
                item[col] = val if val else ""
            else:
                item[col] = ""
        item["_table_index"] = row["index"]
        # Only include rows that have an Order ID or Bill Item ID
        order_id = str(item.get("Order ID (YYMMDD_vendor_gburdell3)", "") or item.get("Order ID", "")).strip()
        bill_item_id = str(item.get("Bill Item ID", "")).strip()
        if order_id or bill_item_id:
            items.append(item)
    return items


def graph_update_order_status(order_id_col_name: str, order_id: str, status: str, columns: list[str]) -> bool:
    """Update the Status column for all rows matching an Order ID in OrderT."""
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        return False

    access_token, drive_id, file_id = creds
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

    # Get rows
    rows_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/OrderT/rows"
    resp = _requests.get(rows_url, headers=headers, timeout=15)
    if resp.status_code != 200:
        return False

    # Find order ID column index and status column index
    order_col_idx = None
    status_col_idx = None
    for i, col in enumerate(columns):
        if "Order ID" in col:
            order_col_idx = i
        if col == "Status":
            status_col_idx = i

    if order_col_idx is None or status_col_idx is None:
        return False

    # Find matching rows and update status
    updated = 0
    for row in resp.json().get("value", []):
        vals = row["values"][0]
        row_order_id = str(vals[order_col_idx]).strip() if order_col_idx < len(vals) and vals[order_col_idx] else ""
        if row_order_id == order_id:
            # Update status cell: OrderT header is row 2, data starts row 3
            sheet_row = row["index"] + 3  # +3 because row 1=TOTALS, row 2=header, index is 0-based
            col_letter = chr(65 + status_col_idx)
            url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/worksheets('Ordering')/range(address='{col_letter}{sheet_row}')"
            r = _requests.patch(url, headers=headers, json={"values": [[status]]}, timeout=10)
            if r.status_code == 200:
                updated += 1

    return updated > 0


def graph_apply_spacer_formatting() -> bool:
    """
    Apply pink conditional formatting to spacer rows on the Ordering sheet via Graph API.
    Spacer rows have an Order ID in column A but no Bill Item ID in column B.
    Formula: =AND($A3<>"",$B3="")
    
    Note: The Graph API ConditionalFormat endpoint is available but has limitations.
    This function creates a custom conditional format on the OrderT data range.
    """
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        print("[graph] No credentials for formatting")
        return False

    access_token, drive_id, file_id = creds
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    # Get the used range on the Ordering sheet to know the extent
    range_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/worksheets('Ordering')/usedRange"
    resp = _requests.get(range_url, headers=headers, timeout=10)
    if resp.status_code != 200:
        print(f"[graph] Failed to get used range: {resp.status_code}")
        return False

    used_range = resp.json().get("address", "Ordering!A1:R100")
    # Extract the row count from the used range
    # Format is like "Ordering!A1:R50"
    import re
    match = re.search(r':([A-Z]+)(\d+)', used_range)
    last_col = match.group(1) if match else "R"
    last_row = int(match.group(2)) if match else 100

    # Apply conditional formatting to rows 3 through last_row (data starts row 3)
    # The range should cover the full data area
    format_range = f"A3:{last_col}{last_row}"

    # Create conditional format via Graph API
    cf_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/worksheets('Ordering')/range(address='{format_range}')/conditionalFormats/add"

    payload = {
        "type": "custom",
        "rule": {
            "formula": '=AND($A3<>"",$B3="")',
            "format": {
                "fill": {
                    "color": "#FFB6C1"  # Light pink
                }
            }
        }
    }

    resp = _requests.post(cf_url, headers=headers, json=payload, timeout=15)

    if resp.status_code in (200, 201):
        print(f"[graph] ✅ Applied pink conditional formatting to spacer rows on Ordering!{format_range}")
        return True
    else:
        print(f"[graph] ❌ Failed to apply conditional formatting: {resp.status_code} {resp.text[:200]}")
        return False


def delete_item(item_id: str) -> bool:
    """Clear an item row by Bill Item ID via Graph API (preserves table structure)."""
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        return False

    access_token, drive_id, file_id = creds
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

    # Find the row with this Bill Item ID
    rows_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/BillsT/rows"
    resp = _requests.get(rows_url, headers=headers, timeout=15)
    if resp.status_code != 200:
        return False

    target_idx = None
    for row in resp.json().get("value", []):
        vals = row["values"][0]
        row_id = str(vals[0]).strip() if vals[0] else ""
        if row_id and row_id == str(item_id).strip():
            target_idx = row["index"]
            break

    if target_idx is None:
        # Fallback: try matching by index directly (item_id might be the table index)
        try:
            idx = int(item_id)
            if 0 <= idx < len(resp.json().get("value", [])):
                target_idx = idx
        except (ValueError, TypeError):
            pass

    if target_idx is None:
        print(f"[graph] ⚠️ Could not find item with ID {item_id}")
        return False

    # Clear the row (columns B through J and L through P, skip A and K which have formulas)
    sheet_row = target_idx + 2
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/worksheets('Bills')/range(address='B{sheet_row}:J{sheet_row}')"
    resp = _requests.patch(url, headers=headers, json={"values": [[""] * 9]}, timeout=10)
    url2 = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/worksheets('Bills')/range(address='L{sheet_row}:P{sheet_row}')"
    resp2 = _requests.patch(url2, headers=headers, json={"values": [[""] * 5]}, timeout=10)

    if resp.status_code == 200 and resp2.status_code == 200:
        # Reset cache
        global _cached_items, _cached_items_time
        _cached_items = []
        _cached_items_time = 0
        print(f"[graph] ✅ Cleared row {sheet_row} (item ID {item_id})")
        return True

    return False
