"""
spreadsheet_utils.py - Robust spreadsheet loading, column normalization, and pre-flight validation.

Protects against common user editing variations in FY27_Bills_Budget.xlsx:
- Header row offsets (header on row 0, 1, or 2)
- Case/space variations in column names ("Bill Item ID", "bill item id", "Bill No.", "Bill Number")
- Floating point ID formatting (376851.0 -> "376851")
- Missing links, duplicate IDs, or unlinked Order rows
"""

from __future__ import annotations

import os
import re
import pandas as pd
import openpyxl


COLUMN_ALIASES = {
    "bill_item_id": ["bill item id", "bill_item_id", "item id", "id", "bill item #"],
    "bill_no": ["bill no.", "bill no", "bill_no", "bill #", "bill number", "bill_number"],
    "bill_title": ["bill title", "bill_title", "bill name", "bill_name", "title"],
    "item_name": ["item name", "item_name", "item", "description / item"],
    "budget_section": ["budget section", "budget_section", "section", "category"],
    "vendor": ["vendor", "supplier", "merchant"],
    "description": ["description", "desc", "details", "notes"],
    "quantity": ["quantity", "qty", "count"],
    "cost": ["cost", "unit cost", "cost ($)", "price", "allocation"],
    "total_cost": ["total cost", "total_cost", "total", "total ($)"],
    "link": ["link", "url", "product link", "item link", "product url"],
    "status": ["status", "state", "item status"],
    "order_id": ["order id", "order_id", "order #", "order number", "order_id (yymmdd_vendor_gburdell3)"],
}


def clean_str(val) -> str:
    """Safely convert value to string, handling floats like 376851.0 -> '376851'."""
    if val is None or pd.isna(val):
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2].strip()
    if s.lower() in ("nan", "none", "null"):
        return ""
    return s


def clean_id(val) -> str:
    """Sanitize ID fields (Bill Item ID, Bill No.)."""
    return clean_str(val)


def safe_float(val, default: float = 0.0) -> float:
    """Parse string or number into float safely."""
    if val is None or pd.isna(val):
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).replace("$", "").replace(",", "").strip()
        return float(s)
    except (ValueError, TypeError):
        return default


def safe_int(val, default: int = 1) -> int:
    """Parse string or float into integer safely."""
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def find_sheet_name(excel_file: pd.ExcelFile | str | openpyxl.Workbook, candidate_names: list[str]) -> str | None:
    """Find sheet name ignoring case and slight differences."""
    if isinstance(excel_file, str):
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            existing_sheets = wb.sheetnames
            wb.close()
        except Exception:
            existing_sheets = []
    elif hasattr(excel_file, "sheetnames"):
        existing_sheets = excel_file.sheetnames
    elif hasattr(excel_file, "sheet_names"):
        existing_sheets = excel_file.sheet_names
    else:
        existing_sheets = []

    norm_existing = {s.lower().strip(): s for s in existing_sheets}
    for cand in candidate_names:
        cand_clean = cand.lower().strip()
        if cand_clean in norm_existing:
            return norm_existing[cand_clean]
    for cand in candidate_names:
        cand_clean = cand.lower().strip()
        for norm_name, original_name in norm_existing.items():
            if cand_clean in norm_name or norm_name in cand_clean:
                return original_name
    return None


def get_col_val(row_dict: dict, canonical_key: str, default: str = "") -> str:
    """Extract a row value using flexible column alias matching."""
    aliases = COLUMN_ALIASES.get(canonical_key, [canonical_key])
    norm_row = {str(k).lower().strip(): v for k, v in row_dict.items()}
    for alias in aliases:
        if alias in norm_row:
            return clean_str(norm_row[alias])
    return default


def read_sheet_robust(excel_file: pd.ExcelFile | str | openpyxl.Workbook, sheet_candidates: list[str], max_header_scan: int = 10) -> pd.DataFrame:
    """
    Read an Excel sheet robustly by searching the first max_header_scan rows
    for the true header row containing key column names.
    """
    sheet_name = find_sheet_name(excel_file, sheet_candidates)
    if not sheet_name:
        return pd.DataFrame()

    if isinstance(excel_file, openpyxl.Workbook):
        ws = excel_file[sheet_name]
        data = list(ws.iter_rows(values_only=True))
        df_raw = pd.DataFrame(data)
    else:
        df_raw = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, engine="openpyxl")

    if df_raw.empty:
        return pd.DataFrame()

    header_row_idx = 0
    best_match_count = 0

    all_alias_strings = set()
    for aliases in COLUMN_ALIASES.values():
        all_alias_strings.update(aliases)

    for row_idx in range(min(max_header_scan, len(df_raw))):
        row_vals = [str(v).lower().strip() for v in df_raw.iloc[row_idx].dropna()]
        matches = sum(1 for v in row_vals if any(alias in v for alias in all_alias_strings))
        if matches > best_match_count:
            best_match_count = matches
            header_row_idx = row_idx

    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row_idx).astype(object).fillna("")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def validate_budget_spreadsheet(xlsx_path: str) -> dict:
    """
    Run a full diagnostic health check on FY27_Bills_Budget.xlsx.
    Returns dict with 'valid' (bool), 'errors' (list of str), 'warnings' (list of str), 'summary' (str).
    """
    results = {"valid": True, "errors": [], "warnings": [], "summary": ""}

    if not os.path.exists(xlsx_path):
        results["valid"] = False
        results["errors"].append(f"Spreadsheet file not found at: {xlsx_path}")
        results["summary"] = "CRITICAL: Excel file missing."
        return results

    try:
        excel_file = pd.ExcelFile(xlsx_path)
    except Exception as ex:
        results["valid"] = False
        results["errors"].append(f"Failed to open Excel file: {ex}")
        results["summary"] = "CRITICAL: Excel file corrupt or unreadable."
        return results

    # 1. Validate Bills Sheet
    bills_sheet = find_sheet_name(excel_file, ["Bills", "Bill", "Budget"])
    if not bills_sheet:
        results["valid"] = False
        results["errors"].append("Missing required 'Bills' sheet in workbook.")
    else:
        df_bills = read_sheet_robust(excel_file, [bills_sheet])
        if df_bills.empty:
            results["warnings"].append(f"'${bills_sheet}' sheet contains no data rows.")
        else:
            seen_item_ids = {}
            for idx, row in df_bills.iterrows():
                excel_line = idx + 2  # 1-indexed Excel row after header
                r_dict = row.to_dict()
                b_id = get_col_val(r_dict, "bill_item_id")
                item_name = get_col_val(r_dict, "item_name")
                bill_no = get_col_val(r_dict, "bill_no")
                cost = safe_float(r_dict.get("Cost", r_dict.get("Allocation", 0)))
                link = get_col_val(r_dict, "link")

                if not item_name and not b_id:
                    continue  # skip empty separator rows

                if b_id:
                    if b_id in seen_item_ids:
                        results["errors"].append(
                            f"Duplicate Bill Item ID '{b_id}' on row {excel_line} (previously seen on row {seen_item_ids[b_id]})."
                        )
                        results["valid"] = False
                    else:
                        seen_item_ids[b_id] = excel_line

                if not item_name:
                    results["warnings"].append(f"Row {excel_line} has Bill Item ID '{b_id}' but missing Item Name.")
                if not bill_no:
                    results["warnings"].append(f"Row {excel_line} ('{item_name}') missing Bill No.")
                if cost <= 0:
                    results["warnings"].append(f"Row {excel_line} ('{item_name}') has cost <= $0.00 (${cost:.2f}).")
                if link and not link.startswith("http"):
                    results["warnings"].append(f"Row {excel_line} ('{item_name}') has non-standard link: '{link}'.")

    # 2. Validate Ordering Sheet
    ordering_sheet = find_sheet_name(excel_file, ["Ordering", "Orders", "OrderT"])
    if ordering_sheet:
        df_orders = read_sheet_robust(excel_file, [ordering_sheet])
        if not df_orders.empty:
            bills_sheet_name = find_sheet_name(excel_file, ["Bills", "Bill", "Budget"])
            df_bills = read_sheet_robust(excel_file, [bills_sheet_name]) if bills_sheet_name else pd.DataFrame()
            known_b_ids = set()
            if not df_bills.empty:
                for _, r in df_bills.iterrows():
                    bid = get_col_val(r.to_dict(), "bill_item_id")
                    if bid:
                        known_b_ids.add(bid)

            for idx, row in df_orders.iterrows():
                excel_line = idx + 3  # Excel row index
                r_dict = row.to_dict()
                order_id = get_col_val(r_dict, "order_id")
                b_id = get_col_val(r_dict, "bill_item_id")
                item_name = get_col_val(r_dict, "item_name")

                if not order_id or order_id.startswith("Order ") or not (b_id or item_name):
                    continue

                if b_id and known_b_ids and b_id not in known_b_ids:
                    results["warnings"].append(
                        f"Ordering sheet row {excel_line} (Order: '{order_id}') references Bill Item ID '{b_id}' which does not exist in Bills sheet."
                    )

    error_cnt = len(results["errors"])
    warn_cnt = len(results["warnings"])
    if error_cnt == 0 and warn_cnt == 0:
        results["summary"] = "✅ Spreadsheet passed all diagnostic health checks cleanly!"
    elif error_cnt == 0:
        results["summary"] = f"⚠️ Spreadsheet passed with {warn_cnt} warning(s)."
    else:
        results["summary"] = f"❌ Spreadsheet failed validation with {error_cnt} error(s) and {warn_cnt} warning(s)."

    return results
