"""
xlsx_manager.py - Read/write FY27_Bills_Budget.xlsx with rclone sync.

Pull before read, push after write. The xlsx on SharePoint is the source of truth.
"""

import os
import subprocess
import threading
from pathlib import Path
from openpyxl import load_workbook

# Config from environment
RCLONE_REMOTE = os.environ.get(
    "RCLONE_REMOTE",
    "onedrive:OPS-1 Operations/FY27 Finances/FY27_Bills_Budget.xlsx",
)
LOCAL_XLSX = os.environ.get("LOCAL_XLSX_PATH", os.path.expanduser("~/mrg/finance/FY27_Bills_Budget.xlsx"))
SHEET_NAME = os.environ.get("XLSX_SHEET_NAME", "Bills")
QUEUE_SHEET_NAME = os.environ.get("XLSX_QUEUE_SHEET_NAME", "Test")

# File lock to prevent concurrent reads/writes
_lock = threading.Lock()

# Cache: only pull from SharePoint every N seconds
_last_pull_time = 0
PULL_INTERVAL = int(os.environ.get("PULL_INTERVAL_SECONDS", "300"))  # 5 minutes

# Cached data
_cached_items: list[dict] = []
_cached_items_time = 0
ITEMS_CACHE_TTL = 60  # seconds

# Column mapping (xlsx columns in the Bills sheet)
COLUMNS = [
    "Bill Item ID",
    "Bill No.",
    "Bill Title",
    "Item Name",
    "Status",
    "Budget Section",
    "Vendor",
    "Description",
    "Quantity",
    "Cost",
    "Total Cost",
    "Link",
    "File URL",
    "Person Requesting",
]


def _run_rclone(args: list[str]) -> bool:
    """Run an rclone command. Returns True on success."""
    try:
        result = subprocess.run(
            ["rclone"] + args,
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            print(f"[rclone] Error: {result.stderr.strip()}")
            return False
        return True
    except FileNotFoundError:
        print("[rclone] rclone not found - skipping sync")
        return False
    except subprocess.TimeoutExpired:
        print("[rclone] Timeout during sync")
        return False


def sync_pull() -> bool:
    """Pull latest xlsx from SharePoint. Skips if pulled recently (within PULL_INTERVAL)."""
    global _last_pull_time
    import time as _time

    now = _time.time()
    if now - _last_pull_time < PULL_INTERVAL:
        return True  # Use cached local copy

    local_dir = str(Path(LOCAL_XLSX).parent)
    os.makedirs(local_dir, exist_ok=True)
    result = _run_rclone(["copy", RCLONE_REMOTE, local_dir])
    if result:
        _last_pull_time = now
    return result


def _get_graph_token() -> tuple[str, str, str] | None:
    """Read access token, drive_id, and file_id from rclone config / cache."""
    import configparser
    import json as _json

    rclone_conf = os.path.expanduser("~/.config/rclone/rclone.conf")
    if not os.path.exists(rclone_conf):
        return None

    config = configparser.ConfigParser()
    config.read(rclone_conf)

    if "onedrive" not in config:
        return None

    try:
        token_str = config["onedrive"]["token"]
        token = _json.loads(token_str)
        drive_id = config["onedrive"]["drive_id"]
        access_token = token["access_token"]
    except (KeyError, _json.JSONDecodeError):
        return None

    # Get file ID (cache it after first lookup)
    file_id = os.environ.get("_GRAPH_FILE_ID", "")
    if not file_id:
        import requests as _requests
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/OPS-1 Operations/FY27 Finances/FY27_Bills_Budget.xlsx"
        resp = _requests.get(url, headers={"Authorization": f"Bearer {access_token}"}, timeout=10)
        if resp.status_code == 200:
            file_id = resp.json()["id"]
            os.environ["_GRAPH_FILE_ID"] = file_id
        else:
            return None

    return access_token, drive_id, file_id


def graph_add_row(sheet_table: str, row_values: list) -> bool:
    """Add a row to a table via Graph API Excel workbook endpoint.
    Inserts after the last non-empty row to avoid gaps.
    """
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        print("[graph] No credentials - skipping")
        return False

    access_token, drive_id, file_id = creds
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    # Find the last non-empty row index
    rows_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/{sheet_table}/rows"
    rows_resp = _requests.get(rows_url, headers=headers, timeout=15)

    insert_index = None
    if rows_resp.status_code == 200:
        rows = rows_resp.json().get("value", [])
        # Find last row with data (check Item Name column, index 3 for BillsT, 3 for TestTable)
        last_data_index = -1
        for row in rows:
            vals = row["values"][0] if row.get("values") else []
            # Check if any cell has data
            if any(str(v).strip() for v in vals if v):
                last_data_index = row["index"]
        if last_data_index >= 0:
            insert_index = last_data_index + 1

    # Insert at specific index or append
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/{sheet_table}/rows/add"
    payload = {"values": [row_values]}
    if insert_index is not None:
        payload["index"] = insert_index

    resp = _requests.post(url, headers=headers, json=payload, timeout=15)

    if resp.status_code in (200, 201):
        print(f"[graph] ✅ Row added to {sheet_table} at index {insert_index}")
        return True
    else:
        print(f"[graph] ❌ Failed to add row: {resp.status_code} {resp.text[:150]}")
        return False


def graph_get_table_columns(sheet_table: str) -> list[str]:
    """Get column names for a table via Graph API."""
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        return []

    access_token, drive_id, file_id = creds

    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/{sheet_table}/columns"
    resp = _requests.get(
        url,
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=10,
    )

    if resp.status_code == 200:
        return [c["name"] for c in resp.json()["value"]]
    return []


def sync_push() -> bool:
    """No-op — writes now go directly via Graph API."""
    return True


def _push_async():
    """No-op — writes are immediate via Graph API."""
    pass


# Titles to skip (same as automation.py)
SKIP_TITLE_PREFIXES = ("nan", "request", "liquid", "misc")


def read_items() -> list[dict]:
    """
    Read all items from the Bills sheet.
    Pulls latest from SharePoint first.
    Auto-detects header row by scanning for 'Item Name' column.
    Skips non-bill items (Liquid, Misc, etc.) same as automation.py.
    Returns list of dicts with normalized keys.
    """
    with _lock:
        sync_pull()

        if not os.path.exists(LOCAL_XLSX):
            return []

        wb = load_workbook(LOCAL_XLSX, read_only=True, data_only=True)
        ws = wb[SHEET_NAME]

        # Auto-detect header row by finding the row containing "Item Name"
        header_row = None
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_values = [str(cell).strip() if cell else "" for cell in row]
            if "Item Name" in row_values:
                header_row = row_idx
                headers = row_values
                break

        if header_row is None:
            wb.close()
            return []

        items = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(row):
                continue
            item = {}
            for i, col in enumerate(headers):
                if not col:
                    continue
                if i < len(row):
                    val = row[i]
                    # Normalize whitespace in strings
                    if isinstance(val, str):
                        val = " ".join(val.split())
                    item[col] = val if val is not None else ""
                else:
                    item[col] = ""
            # Only include rows that have an Item Name
            if not item.get("Item Name"):
                continue
            # Skip non-bill items (same filter as automation.py)
            bill_title = str(item.get("Bill Title", "")).strip().lower()
            if any(bill_title.startswith(prefix) for prefix in SKIP_TITLE_PREFIXES):
                continue
            # Skip items with negative Bill Item IDs (metadata rows)
            try:
                item_id = float(str(item.get("Bill Item ID", 0)))
                if item_id < 0:
                    continue
            except (ValueError, TypeError):
                pass
            items.append(item)

        wb.close()
        return items


def get_bills() -> list[str]:
    """Get list of unique bill titles (non-empty)."""
    items = read_items()
    titles = set()
    for item in items:
        title = str(item.get("Bill Title", "")).strip()
        if title:
            titles.add(title)
    return sorted(titles)


def get_items_by_bill(bill_title: str) -> list[dict]:
    """Get items for a specific bill."""
    items = read_items()
    return [i for i in items if str(i.get("Bill Title", "")).strip() == bill_title]


def get_backlog_items() -> list[dict]:
    """Get items from the queue (Test sheet)."""
    return read_queue_items()


def _find_row_by_item_id(ws, item_id: str, headers: list[str], header_row: int) -> int | None:
    """Find the row number for a given Bill Item ID."""
    id_col = None
    for i, h in enumerate(headers):
        if h == "Bill Item ID":
            id_col = i
            break
    if id_col is None:
        return None

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if id_col < len(row) and str(row[id_col]).strip() == str(item_id).strip():
            return row_idx
    return None


def _get_headers(ws) -> tuple[list[str], int]:
    """Get header row as list of strings and the 1-indexed row number."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_values = [str(cell).strip() if cell else "" for cell in row]
        if "Item Name" in row_values:
            return row_values, row_idx
    # Fallback to row 1
    return [str(cell.value).strip() if cell.value else "" for cell in ws[1]], 1


def _get_next_item_id(ws, headers: list[str], header_row: int) -> int:
    """Get the next available Bill Item ID."""
    id_col = None
    for i, h in enumerate(headers):
        if h == "Bill Item ID":
            id_col = i
            break
    if id_col is None:
        return 1

    max_id = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if id_col < len(row) and row[id_col]:
            try:
                val = int(float(str(row[id_col])))
                max_id = max(max_id, val)
            except (ValueError, TypeError):
                pass
    return max_id + 1


def add_item(item_data: dict) -> bool:
    """
    Add a new item to the QUEUE (TestTable on Test sheet) via Graph API.
    Also writes locally for immediate display.
    """
    # Get TestTable columns
    columns = graph_get_table_columns("TestTable")
    if not columns:
        # Fallback: use known columns
        columns = ["Bill Item ID", "Bill No.", "Bill Title", "Item Name",
                   "Budget Section", "Quantity", "Cost", "Vendor", "Description", "Link", "Column1"]

    # Build row values in column order
    row_values = []
    for col in columns:
        val = item_data.get(col, "")
        row_values.append(val if val else "")

    # Push to SharePoint via Graph API
    success = graph_add_row("TestTable", row_values)

    # Also write locally for immediate read-back
    with _lock:
        if os.path.exists(LOCAL_XLSX):
            try:
                wb = load_workbook(LOCAL_XLSX)
                ws = wb[QUEUE_SHEET_NAME]
                headers, header_row = _get_headers(ws)
                local_row = []
                for h in headers:
                    local_row.append(item_data.get(h, ""))
                ws.append(local_row)
                wb.save(LOCAL_XLSX)
                wb.close()
            except Exception as e:
                print(f"[local] Write failed: {e}")

    return success


def read_queue_items() -> list[dict]:
    """
    Read all items from the Queue (TestTable) via Graph API.
    Falls back to local xlsx if Graph API unavailable.
    """
    import requests as _requests

    creds = _get_graph_token()
    if creds:
        access_token, drive_id, file_id = creds
        headers = {"Authorization": f"Bearer {access_token}"}

        # Get columns
        cols_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/TestTable/columns"
        cols_resp = _requests.get(cols_url, headers=headers, timeout=10)

        # Get rows
        rows_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/TestTable/rows"
        rows_resp = _requests.get(rows_url, headers=headers, timeout=10)

        if cols_resp.status_code == 200 and rows_resp.status_code == 200:
            columns = [c["name"] for c in cols_resp.json()["value"]]
            items = []
            for row in rows_resp.json().get("value", []):
                values = row["values"][0] if row.get("values") else []
                item = {}
                for i, col in enumerate(columns):
                    if i < len(values):
                        val = values[i]
                        if isinstance(val, str):
                            val = " ".join(val.split())
                        item[col] = val if val else ""
                    else:
                        item[col] = ""
                if item.get("Item Name"):
                    item["_table_index"] = row["index"]
                    items.append(item)
            return items

    # Fallback to local xlsx
    with _lock:
        sync_pull()
        if not os.path.exists(LOCAL_XLSX):
            return []

        wb = load_workbook(LOCAL_XLSX, read_only=True, data_only=True)
        ws = wb[QUEUE_SHEET_NAME]
        headers_list, header_row = _get_headers(ws)

        items = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(row):
                continue
            item = {}
            for i, col in enumerate(headers_list):
                if not col:
                    continue
                if i < len(row):
                    val = row[i]
                    if isinstance(val, str):
                        val = " ".join(val.split())
                    item[col] = val if val is not None else ""
                else:
                    item[col] = ""
            if item.get("Item Name"):
                item["_table_index"] = row_idx - header_row - 1  # 0-indexed for Graph API
                items.append(item)

        wb.close()
        return items


def delete_queue_item(row_idx: int) -> bool:
    """Delete an item from the queue (Test sheet) by row index."""
    with _lock:
        if not os.path.exists(LOCAL_XLSX):
            return False

        wb = load_workbook(LOCAL_XLSX)
        ws = wb[QUEUE_SHEET_NAME]
        ws.delete_rows(row_idx)
        wb.save(LOCAL_XLSX)
        wb.close()
        return True


def move_to_bill(queue_items: list[dict], bill_title: str) -> int:
    """
    Move items from the Queue (TestTable) to the Bills table (BillsT) via Graph API.
    - Adds each item as a new row on BillsT with the given Bill Title
    - Deletes them from TestTable
    - Returns number of items successfully moved.
    """
    import requests as _requests

    creds = _get_graph_token()
    if not creds:
        print("[graph] No credentials")
        return 0

    access_token, drive_id, file_id = creds
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    # Get BillsT columns
    bills_columns = graph_get_table_columns("BillsT")
    if not bills_columns:
        bills_columns = ["Bill Item ID", "Bill No.", "Bill Title", "Item Name", "Status",
                         "Budget Section", "Vendor", "Description", "Quantity", "Cost",
                         "Total Cost", "Link", "File URL", "Person Requesting", "Remaining Allocation", "Column1"]

    # Get next Bill Item ID from BillsT
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/BillsT/rows"
    resp = _requests.get(url, headers=headers, timeout=15)
    max_id = 0
    if resp.status_code == 200:
        for row in resp.json().get("value", []):
            try:
                val = int(float(str(row["values"][0][0])))
                if val > max_id:
                    max_id = val
            except (ValueError, TypeError, IndexError):
                pass
    next_id = max_id + 1

    moved = 0
    rows_to_delete = []

    for item in queue_items:
        # Build row for BillsT
        item_data = dict(item)
        item_data["Bill Item ID"] = next_id
        item_data["Bill Title"] = bill_title
        item_data["Status"] = "Bill Requested"

        # Calculate Total Cost
        try:
            qty = float(item_data.get("Quantity", 1) or 1)
            cost = float(str(item_data.get("Cost", 0)).replace("$", "").replace(",", "") or 0)
            item_data["Total Cost"] = qty * cost
        except (ValueError, TypeError):
            item_data["Total Cost"] = 0

        row_values = []
        for col in bills_columns:
            val = item_data.get(col, "")
            row_values.append(val if val else "")

        # Add to BillsT
        success = graph_add_row("BillsT", row_values)
        if success:
            moved += 1
            next_id += 1
            # Track row index for deletion from TestTable
            if "_table_index" in item:
                rows_to_delete.append(item["_table_index"])

    # Delete from TestTable (in reverse order so indices don't shift)
    rows_to_delete.sort(reverse=True)
    for idx in rows_to_delete:
        del_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/workbook/tables/TestTable/rows/itemAt(index={idx})"
        resp = _requests.delete(del_url, headers=headers, timeout=10)
        if resp.status_code == 204:
            print(f"[graph] Deleted queue row index {idx}")
        else:
            print(f"[graph] Failed to delete row {idx}: {resp.status_code}")

    # Also sync local file
    sync_pull()

    return moved


def update_item(item_id: str, updates: dict) -> bool:
    """
    Update an existing item by Bill Item ID.
    Only modifies specified fields.
    """
    with _lock:
        sync_pull()

        if not os.path.exists(LOCAL_XLSX):
            return False

        wb = load_workbook(LOCAL_XLSX)
        ws = wb[SHEET_NAME]
        headers, header_row = _get_headers(ws)

        row_idx = _find_row_by_item_id(ws, item_id, headers, header_row)
        if row_idx is None:
            wb.close()
            return False

        for key, value in updates.items():
            if key in headers:
                col_idx = headers.index(key) + 1  # openpyxl is 1-indexed
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Recalculate Total Cost if qty or cost changed
        if "Quantity" in updates or "Cost" in updates:
            qty_col = headers.index("Quantity") + 1 if "Quantity" in headers else None
            cost_col = headers.index("Cost") + 1 if "Cost" in headers else None
            total_col = headers.index("Total Cost") + 1 if "Total Cost" in headers else None

            if qty_col and cost_col and total_col:
                qty = ws.cell(row=row_idx, column=qty_col).value or 1
                cost = ws.cell(row=row_idx, column=cost_col).value or 0
                try:
                    cost_val = float(str(cost).replace("$", "").replace(",", ""))
                    qty_val = float(qty)
                    ws.cell(row=row_idx, column=total_col, value=qty_val * cost_val)
                except (ValueError, TypeError):
                    pass

        wb.save(LOCAL_XLSX)
        wb.close()

        _push_async()
        return True


def delete_item(item_id: str) -> bool:
    """Delete an item row by Bill Item ID."""
    with _lock:
        sync_pull()

        if not os.path.exists(LOCAL_XLSX):
            return False

        wb = load_workbook(LOCAL_XLSX)
        ws = wb[SHEET_NAME]
        headers, header_row = _get_headers(ws)

        row_idx = _find_row_by_item_id(ws, item_id, headers, header_row)
        if row_idx is None:
            wb.close()
            return False

        ws.delete_rows(row_idx)
        wb.save(LOCAL_XLSX)
        wb.close()

        _push_async()
        return True
