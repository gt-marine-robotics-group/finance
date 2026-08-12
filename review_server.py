"""
Local server for review.html to save price edits directly to the SharePoint-synced xlsx.

Start this before opening review.html:
    python review_server.py

It runs on http://localhost:8321 and accepts price updates from the browser.
"""

import os
import sys
import json
import webbrowser
from http.server import HTTPServer, SimpleHTTPRequestHandler
import openpyxl

XLSX_PATH = os.environ.get("FINANCE_XLSX_PATH", os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-GeorgiaInstituteofTechnology/"
    "Documents - Marine Robotics Group/OPS-1 Operations/FY27 Finances/FY27_Bills_Budget.xlsx"
))
SHEET_NAME = "Bills"
PORT = 8321


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


class ReviewHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=SCRIPT_DIR, **kwargs)

    def _normalize_path(self):
        clean_path = self.path.split("?")[0].rstrip("/")
        if (
            not clean_path
            or clean_path in ("", "/review", "/review.html")
            or clean_path.startswith("/orders/review")
            or clean_path.startswith("/review")
        ):
            self.path = "/review.html"
            return

        requested_file = os.path.join(SCRIPT_DIR, self.path.lstrip("/"))
        if not os.path.isfile(requested_file):
            self.path = "/review.html"

    def do_GET(self):
        self._normalize_path()
        target_file = os.path.join(SCRIPT_DIR, self.path.lstrip("/"))
        if self.path == "/review.html" and not os.path.isfile(target_file):
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            fallback_html = """<!DOCTYPE html>
<html>
<head><title>Side-by-Side Review</title><meta http-equiv="refresh" content="2"></head>
<body style="font-family:sans-serif; background:#f8fafc; padding:40px; text-align:center; color:#1e293b;">
    <h2>📋 Preparing Side-by-Side Review Page...</h2>
    <p>Please wait a moment while screenshots and prices are prepared.</p>
</body>
</html>"""
            self.wfile.write(fallback_html.encode("utf-8"))
            return
        return super().do_GET()

    def do_HEAD(self):
        self._normalize_path()
        return super().do_HEAD()

    def do_POST(self):
        if self.path == "/save-prices":
            content_length = int(self.headers["Content-Length"])
            body = self.rfile.read(content_length)
            data = json.loads(body)

            try:
                updates = data.get("prices", [])
                if not updates:
                    self._respond(400, {"error": "No prices provided"})
                    return

                # Open xlsx and update
                wb = openpyxl.load_workbook(XLSX_PATH)
                ws = wb[SHEET_NAME]

                # Find column indices
                headers = [cell.value for cell in ws[1]]
                name_col = None
                cost_col = None
                for i, h in enumerate(headers):
                    if h and str(h).strip() == "Item Name":
                        name_col = i
                    if h and str(h).strip() == "Cost":
                        cost_col = i

                if name_col is None or cost_col is None:
                    self._respond(500, {"error": "Could not find Item Name or Cost column"})
                    return

                updated = []
                for update in updates:
                    item_name = update["item_name"].strip()
                    new_price = update["price"]

                    # Find the row
                    for row in ws.iter_rows(min_row=2):
                        cell_name = row[name_col].value
                        if cell_name and str(cell_name).strip() == item_name:
                            row[cost_col].value = new_price
                            updated.append(item_name)
                            break

                wb.save(XLSX_PATH)
                print(f"✅ Updated {len(updated)} prices in spreadsheet")
                for name in updated:
                    print(f"   • {name}")

                # Touch the file to trigger OneDrive sync detection
                os.utime(XLSX_PATH, None)

                self._respond(200, {"updated": updated, "count": len(updated)})

            except Exception as e:
                print(f"❌ Error: {e}")
                self._respond(500, {"error": str(e)})

        else:
            self._respond(404, {"error": "Not found"})

    def do_OPTIONS(self):
        """Handle CORS preflight."""
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def _respond(self, code, data):
        self.send_response(code)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data).encode())

    def log_message(self, format, *args):
        # Only log POST requests, not static file requests
        if "POST" in str(args):
            super().log_message(format, *args)


class ReusableHTTPServer(HTTPServer):
    allow_reuse_address = True


def launch_review_server_and_browser(html_path="review.html", port=PORT):
    """Simple, 100% bulletproof review server & browser launcher."""
    import time
    import urllib.request
    import threading
    import subprocess
    import webbrowser

    # Ensure html_path is absolute
    if not os.path.isabs(html_path):
        html_path = os.path.join(SCRIPT_DIR, html_path)

    # Start HTTP server thread in background for Excel POST syncing if needed
    for test_port in (port, 8322, 8323):
        url = f"http://127.0.0.1:{test_port}"
        try:
            urllib.request.urlopen(url, timeout=0.3)
            break
        except Exception:
            try:
                server = ReusableHTTPServer(("0.0.0.0", test_port), ReviewHandler)
                server_thread = threading.Thread(target=server.serve_forever, daemon=True)
                server_thread.start()
                time.sleep(0.2)
                print(f"  🚀 Background sync server listening on {url}")
                break
            except Exception:
                continue

    file_uri = f"file://{os.path.abspath(html_path)}"
    print(f"  🌐 Opening Review Page: {file_uri}")

    if sys.platform == "darwin":
        subprocess.run(["open", os.path.abspath(html_path)], check=False)
    else:
        webbrowser.open(file_uri)


if __name__ == "__main__":
    print(f"📝 Review server running on http://127.0.0.1:{PORT}")
    print(f"   Saving to: {XLSX_PATH}")
    url = f"http://127.0.0.1:{PORT}"
    try:
        webbrowser.open(url)
    except Exception:
        pass
    try:
        server = ReusableHTTPServer(("0.0.0.0", PORT), ReviewHandler)
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n👋 Server stopped")

